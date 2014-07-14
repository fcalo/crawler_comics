#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys, os, urllib2, urllib, cookielib, re, gzip, StringIO
from lxml import etree
from urllib import quote
from datetime import datetime, timedelta
from datetime import datetime, timedelta
import time, logging, logging.handlers
from pprint import pprint
from binascii import crc32
import traceback
from openpyxl import load_workbook

from utils import *
from crawl import CrawlerComics
from db import DB

class CrawlerComics_2(CrawlerComics):
	def __init__(self, verbose = False, id_task = None, mode = "0"):
		self.verbose = verbose
		
		# 0 -> complete
		# 1 -> only updates and deletes
		self.mode_complete = mode == "0"
		

		self.parser = etree.HTMLParser()
		
		#config
		self.config = {}
		config_file = os.path.join(os.path.dirname(__file__), "crawler_comics_2.conf")
		execfile(config_file, self.config)
		
		#logger
		self.logger = logging.getLogger('CRAWLER')
		hdlr = logging.handlers.TimedRotatingFileHandler(os.path.join(os.path.dirname(__file__), \
		  self.config['log_file'].replace(".log", "%s.log" % id_task)),"d",2)
		hdlr.suffix = "-%s" % id_task if id_task else "%Y-%m-%d-%H-%M"
		formatter = logging.Formatter('%(asctime)s - %(levelname)s: %(message)s')
		hdlr.setFormatter(formatter)
		self.logger.addHandler(hdlr)
		self.logger.setLevel(logging.INFO)
		self.logger.info("[__init__]")
		
			
		self.xpaths = {"name": ['.//table/tr[1]/td/div[1]/text()'],
				"title": ['.//table/tr[1]/td/div[1]/text()'],
				"id": ['.//table/tr[6]/td[2]/text()'],
				"mfgid": ['.//table/tr[6]/td[2]/text()'],
				"extra_field_7": ['.//table/tr[5]/td[2]/text()'],
				"extra_field_11": ['.//table/tr[5]/td[2]/text()'],
				"subcategory": ['.//table/tr[11]/td[2]/text()'],
				"extra_field_5": ['.//table/tr[7]/td[2]/text()'],
				"extra_field_10": ['.//table/tr[2]/td[2]//text()'],
				"extra_field_1": ['.//table/tr[12]/td[2]/text()'],
				"extra_field_4": ['.//table/tr[8]/td[2]/text()'],
				"price2": ['.//table/tr[3]/td[2]/text()'],
				"description": ['.//table/tr[13]/td//text()'],
				"extended_description": ['.//table/tr[13]/td//text()'],
				"image1": ['.//img/@src'],
				"thumbnail": ['.//img/@src'],
				"content": ['.//text()']
			}
		
		
		
		self.category_alias = {"BABEL" : "COMIC EUROPEO"
			, "Babel" : "COMIC EUROPEO"
			, u"BD - Autores Européos" : "COMIC EUROPEO"
			, u"BD - Autores Europeos" : "COMIC EUROPEO"
			, u"BD - AUTORES EUROPEOS" : "COMIC EUROPEO"
			, u"Colección Trazado" : "COMIC INDEPENDIENTE"
			, u"Cómics Clásicos" : "HUMOR"
			, u"Cómics Españoles" : u"COMIC ESPAÑOL"
			, u"Cómics Star Wars" : u"COMIC USA"
			, u"Guías Ilustradas Star Wars" : u"COMIC USA"
			, u"Independientes USA" : u"COMIC USA"
			, u"Novelas Star Wars" : u"COMIC USA"
			}
		self.category_ban = {"MERCHANDISING LOS MUERTOS VIVIENTES":"", "Merchandising Los Muertos Vivientes" : ""}
		
		self.db = DB(self.logger, config_file)
		self.db.init_model()
		
		if not id_task:
			self.id_task = self.db.start_new_task()
		else:
			self.id_task = int(id_task)
			
		#initialite csv
		self.filename_csv = os.path.join(os.path.dirname(__file__), "csv/%s" % self.config['csv_filename'] % self.id_task)
		self.filename_xlsx = os.path.join(os.path.dirname(__file__), "csv/EXTERNAL_%s.xlsx" % self.id_task)
		self.data_external_xml = None
		
		self.print_line(self.config["csv_header"], True)
		self.cj = None
		

	
	def init_metas(self, previous_metas = False):
		self.metas = {"distributor" : self.config['distributor'], "category": "COMICS"
		,"manufacturer" : self.config['distributor'], "tax_code" : "IVL", "extra_field_13": 0 if previous_metas else 2}
		
		
	def download_url(self, url, level = False):
		
		if self.cj is None:
			self.cj = cookielib.CookieJar()
		
		cj = self.cj
		
		opener = urllib2.build_opener(urllib2.HTTPCookieProcessor(cj))

		opener.addheaders = [('User-agent', self.config['user_agent'])]

		urllib2.install_opener(opener)

		authentication_url = self.config['url_login']
		
		post = {'usuari':self.config['login'],
		        'contrasenya':self.config['password'],
		        }


		req = urllib2.Request(self.config['url_login'], urllib.urlencode(post))
		
		downloaded = False
		tries = 0
		while not downloaded:
			try:
				tries += 1
				resp = opener.open(req)
				#~ resp = urllib2.urlopen(req)
				downloaded = True
			except urllib2.URLError as e:
				self.logger.info("[download_url] Error descargando %s - %s" % (url, str(e)))
				if tries == 5 and not level:
					return self.download_url(url, True)
				if tries > 5:
					raise
				else:
					self.logger.info("[download_url] Reintentando ...")
				time.sleep(tries)
			
		foo = resp.read()
		
		url = quote(url.encode("utf-8"),":/&?=")
		
		str_post = "f_titulo=&f_autor=&f_EAN=&f_codigo=&f_coleccion=-1&filtrar_comics=Filtrar"  
		
		if "fondo.php" in url:
			req = urllib2.Request(url, str_post)
		else:
			req = urllib2.Request(url)
		
		
		downloaded = False
		tries = 0
		while not downloaded:
			try:
				tries += 1
				
				resp = opener.open(req)
				
				#~ resp = urllib2.urlopen(req)
				downloaded = True
			except urllib2.URLError as e:
				self.logger.info("[download_url_login] Error descargando %s - %s" % (url, str(e)))
				if tries > 5:
					raise
				else:
					self.logger.info("[download_url_login] Reintentando ...")
				time.sleep(tries)
			
		if 'content-encoding' in resp.headers and resp.headers['content-encoding'] == 'gzip':
			try:
				data = gzip.GzipFile(fileobj = StringIO.StringIO(resp.read())).read()
			except IOError:
				return None
		else:
			data = resp.read()
			
		if "fondo.php" in url and not url.endswith("fondo.php?p=1"):
			req = urllib2.Request(url)
		
			downloaded = False
			tries = 0
			while not downloaded:
				try:
					tries += 1
					
					resp = opener.open(req)
					
					#~ resp = urllib2.urlopen(req)
					downloaded = True
				except urllib2.URLError as e:
					self.logger.info("[download_url_login] Error descargando %s - %s" % (url, str(e)))
					if tries > 5:
						raise
					else:
						self.logger.info("[download_url_login] Reintentando ...")
					time.sleep(tries)
				
			if 'content-encoding' in resp.headers and resp.headers['content-encoding'] == 'gzip':
				try:
					data = gzip.GzipFile(fileobj = StringIO.StringIO(resp.read())).read()
				except IOError:
					return None
			else:
				data = resp.read()
			
		return data
		
	def get_data_from_external_xml(self, _id):
		""" download xlsx with master data """
		
		if not self.data_external_xml:
			self.logger.info("[download_url_login] Descargando maestro web ...")
			#load data
			self.data_external_xml = {}
			
			html = self.download_url(self.config['url_master'])
			link = re.findall('.*?href="([^"]*MAESTRO__LIBRERIAS_WEB.xlsx)".*?', html)[0]
			
			url = ("http://%s/%s" % (self.config['domain'], link)).replace("../","")
			
			f = open(self.filename_xlsx, "w")
			f.write(self.download_url(url))
			f.close()
			
			wb = load_workbook(filename = self.filename_xlsx)
			
			start = False
			finish = False
			row_pos = 1
			while not finish:
				if not start:
					if wb.active.cell(row = row_pos, column = 4).value in ['OK', 'NO DISPONIBLE']:
						start = True
				else:
					if not wb.active.cell(row = row_pos, column = 4).value in ['OK', 'NO DISPONIBLE']:
						finish = True
						
				if start and not finish:
					item = {}
					item['extra_field_7'] = item['extra_field_11'] = wb.active.cell(row = row_pos, column = 1).value
					item['id'] = item['mfgid'] = wb.active.cell(row = row_pos, column = 2).value
					item['title'] = item['name'] = wb.active.cell(row = row_pos, column = 3).value
					item['stock_log'] = wb.active.cell(row = row_pos, column = 4).value
					item['price2'] = wb.active.cell(row = row_pos, column = 5).value
					item['extra_field_10'] = wb.active.cell(row = row_pos, column = 6).value
					item['subcategory'] = wb.active.cell(row = row_pos, column = 7).value
					item['extra_field_4'] = wb.active.cell(row = row_pos, column = 9).value
					item['extra_field_5'] = wb.active.cell(row = row_pos, column = 10).value
					item['extra_field_1'] = "01/01/2008"
					item['description'] = item['extended_description'] = ""
					item['image1'] = item['thumbnail'] = "No_Disponible.gif"
					try:
						item['content'] = " ".join(["" if i is None else i for i in item.values()])
					except:
						self.logger.error(item)
						raise
						
					
					
					self.data_external_xml[item['id']] = item
					
				row_pos += 1

		if _id in self.data_external_xml:
			return self.data_external_xml[_id]
		else:
			return None
			
			
		
		
	def extract_product_campana(self, url):
		""" extract metadata from product reading a support xml"""
		self.logger.info("[extract_product_campana] %s" % url)
		
		n_products = 0
		data_url = self.download_url(url)
		
		
		self.tree = etree.fromstring(data_url, self.parser)
		
		products = self.extracts('//table[@id="comics_campanya"]/tr')
		first = True
		for product in products:
			_id = product.xpath(".//td[2]/text()")[0]
			
			if first:
				
				if self.last_first_id and self.last_first_id == _id:
					#end condition
					return 0
				self.last_first_id = _id
				first = False
					
			#~ print _id
			self.init_metas()
			self.metas = dict(self.metas.items() + self.get_data_from_external_xml(_id).items())
			
			id_product = self.metas['id']
			
			previous_metas = self.db.load_data(id_product)
			
			
			if previous_metas:
				self.metas['extra_field_13'] = 2
				#date in pass?
				now = datetime.now()
				date_created = time.strptime(self.metas['extra_field_1'].strip(), "%d/%m/%Y")
				d_created = datetime(date_created.tm_year, date_created.tm_mon, date_created.tm_mday)
				
				
				if d_created < now and "PROXIMAMENTE" in previous_metas['categories']:
					#not modified but publish date exceeded
					
					#to detect change
					previous_metas['stock'] = self.metas['stock'] + "0"
				else:
					#has been seen before
					
					content = normalize_content(self.metas['content'])
					
					if crc32(content.encode("utf-8")) == previous_metas['crc_content']:
						#no modifications
						self.db.refresh_seen(id_product, self.id_task)
						#ensure images
						if self.config['check_images_without_changes']:
							self.upload_images()
						n_products += 1
						continue
			
			n_products += self.process_metas(id_product, previous_metas = previous_metas)
			
		return n_products
		
	def extract_product(self, url):
		"""extract metadata from product page"""
		
		self.logger.info("[extract_product] %s" % url)
		
		n_products = 0
		data_url = self.download_url(url)
		
		self.tree = etree.fromstring(data_url, self.parser)
		
		proximamente = "novedades.php" in url
		
		products = self.extracts('//table[@id="llistat_comics"]/tr')
		first = True
		for product in products:
			id_product = product.xpath(self.xpaths['id'][0])[0]
			
			if first:
				if self.last_first_id and self.last_first_id == id_product:
					self.logger.info("[extract_product] mismo id que en la página anterior(%s). Terminando" % id_product)
					#end condition
					return 0
				self.last_first_id = id_product
				first = False
			
			self.metas = self.db.load_data(id_product)
			
			previous_metas = {}
			
			if self.metas:
				
				if not proximamente and "PROXIMAMENTE" in self.metas['categories']:
					#not modified but publish date exceeded
					
					#to detect change
					previous_metas['stock'] = self.metas['stock'] + "0"
				else:
					#has been seen before
					content = normalize_content("".join(product.xpath(self.xpaths['content'][0])))
					
					
					if crc32(content.strip().encode("utf-8")) == self.metas['crc_content']:
						#no modifications
						self.db.refresh_seen(id_product, self.id_task)
						#ensure images
						if self.config['check_images_without_changes']:
							self.upload_images()
						
						n_products += 1
						continue
						
					previous_metas['stock'] = self.metas['stock']
					previous_metas['price'] = self.metas['price']
			
			self.init_metas(previous_metas)
			
			for meta, _xpath in self.xpaths.items():
				
				xpath = _xpath[0]
			
				extract = "".join([e for e in product.xpath(xpath) if isinstance(e, basestring)])
				
				if not extract:
					if self.verbose:
						print "\t", meta, _xpath
					continue
				if self.verbose:
					print meta, extract, _xpath
				try:
					self.metas[meta] = extract
				except:
					print "Ha fallado: ", meta, extract, _xpath
					raise
					
				if meta in self.metas:
					self.metas[meta] = self.metas[meta].strip()
					
			n_products +=  self.process_metas(id_product, proximamente, previous_metas)
			
		return n_products
					
	def process_metas(self, id_product, proximamente = False, previous_metas = None):
		""" prepare data """
		
	
		
		
		
		#category validations
		if self.metas['subcategory'] in self.category_alias:
			self.metas['subcategory'] = self.category_alias[self.metas['subcategory']]
		
		
		#category bans
		if self.metas['subcategory'] in self.category_ban:
			return 0
		
		
		
		#categories
		title_collection = get_title_collection(self.metas['title'], self.metas['category'], self.metas['manufacturer'])
		manufacturer = self.metas['manufacturer'] if self.metas['manufacturer'] else "VARIOS"
		
		if not proximamente: 
			
			#CATEGORIA_PRINCIPAL@CATEGORIA_PRINCIPAL/SUBCATEGORIA@CATEGORIA_PRINCIPAL/SUBCATEGORIA/EDITORIAL@CATEGORIA_PRINCIPAL/SUBCATEGORIA/EDITORIAL/TITULO -(menos ó sin) NUMERO COLECCION
			self.metas['categories'] = "%s@%s/%s@%s/%s/%s@%s/%s/%s/%s" % \
			  (self.metas['category'], self.metas['category'], self.metas['subcategory'], \
			  self.metas['category'], self.metas['subcategory'], manufacturer, \
			  self.metas['category'], self.metas['subcategory'], manufacturer, \
			  title_collection)
		else:
			#comming
			self.metas['categories'] = "PROXIMAMENTE@PROXIMAMENTE/%s@PROXIMAMENTE/%s/%s@PROXIMAMENTE/%s/%s/%s" % \
			  (self.metas['category'], self.metas['category'], self.metas['subcategory'], \
			  self.metas['category'], self.metas['subcategory'], manufacturer)
			  
		self.metas['categories'] = "@".join([self.normalize_category(c) for c in self.metas['categories'].split("@")])
		
		#price and cost
		if "por confirmar" in self.metas['price2'].lower():
			self.metas['price2'] = "0"
		else:
			#without euro symbol
			self.metas['price2'] = self.metas['price2'].replace(u"\xa0\u20ac","").replace(u"\u20ac","").replace(u"\x80","")
		
		self.metas['cost'] = float(self.metas['price2'].replace(".","").replace(",",".")) * 0.7
			
		self.metas['price'] = float(self.metas['price2'].replace(".","").replace(",",".")) * 0.95
		
		#date
		date_created = time.strptime(self.metas['extra_field_1'].strip(), "%d/%m/%Y")
		self.metas['date_created'] = time.strftime("%m/%d/%Y", date_created)
		
		#descriptions
		def smart_truncate(content, length=100, suffix=''):
			if len(content) <= length:
				return content
			else:
				return ' '.join(content[:length+1].split(' ')[0:-1]) + suffix
		
		def clean_spaces(s):
			s = ' '.join(s.splitlines())
			while "\t" in s:
				s = s.replace("\t", " ")
			while "  " in s:
				s = s.replace("  ", " ")
			
			return s.strip()
					
		if 'description' in self.metas:
			if self.metas['description'].startswith("Sinopsis:"):
				self.metas['description'] = self.metas['description'][9:]
			if self.metas['extended_description'].startswith("Sinopsis:"):
				self.metas['extended_description'] = self.metas['extended_description'][9:]
				
			self.metas['description'] = smart_truncate(clean_spaces(self.metas['description']))
			self.metas['extended_description'] = clean_spaces(self.metas['extended_description'])
			
		
		#stock & instock_message
		self.metas['stock'] = 40 if proximamente else 10
		if self.metas['stock'] == 10:
			try:
				self.metas['stock'] = 10 if self.get_data_from_external_xml(id_product)['stock_log'] == "OK" else 0
			except TypeError:
				pass
			

		self.metas['instock_message'] = "Pre-Reserva" if self.metas['stock'] == 40 \
		  else "Añadir a Lista de Espera" if self.metas['stock'] == 0 \
		  else "En Stock - 3/5 Días"
		  
		for key_image, sufix in {'thumbnail':'_tb', 'image1':''}.items():
			if key_image in self.metas:
				#~ print self.metas[key_image]
				if not "http" in self.metas[key_image]:
					self.metas[key_image] = ("http://%s/%s" % (self.config['domain'], self.metas[key_image])).replace("../","")
				filename = "%s%s.jpg" % (self.metas["id"], sufix)
				if not self.download_img(self.metas[key_image], filename , thumbnail = key_image == "thumbnail" ):
					del self.metas[key_image]
					continue
				
				
				finalname = "%s%s/%s/%s/%s" % (self.config['url_images'], self.metas['category'], self.metas['subcategory'], \
				  self.metas['manufacturer'], filename)
				self.metas[key_image] = self.normalize_path(finalname)


		#homespecial
		now = datetime.now()
		date_created = time.strptime(self.metas['extra_field_1'].strip(), "%d/%m/%Y")
		
		d_created = datetime(date_created.tm_year, date_created.tm_mon, date_created.tm_mday)
		
		self.metas['homespecial'] = 1 if abs((now - d_created).days) <10 else 0
		
		#reward_points
		self.metas['reward_points'] = int(self.metas['price'] * 20 if d_created > now else self.metas['price'] * 10)
		
		
		#keywords & metatags
		keys_keywords = ["category", "subcategory", "manufacturer", "title", "extra_field_10", "extra_field_3"]
		self.metas['keywords'] = ", ".join(self.metas[i].strip() for i in keys_keywords if i in self.metas and len(self.metas[i])>1)
		
		self.metas['extra_field_7'] = "<div>%s</div>" % self.metas['extra_field_7']
		
		def cut_last_comma(s):
			if s[-1] == ",":
				s = s[:-1]
			if len(s) > 1 and s[-2] == ", ":
				s = s[:-2]
			return s
		
		self.metas['keywords'] = cut_last_comma(self.metas['keywords'])
		if 'extra_field_10' in self.metas:
			self.metas['extra_field_10'] = cut_last_comma(self.metas['extra_field_10'])
		
		self.metas['metatags'] = '<META NAME="KEYWORDS" CONTENT="%s">' % self.metas['keywords']
		
		if previous_metas:
			#has been seen already
			if previous_metas['stock'] == self.metas['stock'] and previous_metas['price'] == self.metas['price']:
				#has modifications but not in price or stock. Dont update.
				return 0
				
		#extra_field_11  
		if 'extra_field_11' in self.metas and self.metas['extra_field_11']:
			self.metas['extra_field_11'] = "<div>%s</div>" % self.metas['extra_field_11']
		
		self.metas['price2'] = self.metas['price2'].replace(",", ".")
		
		self.metas['content'] =  normalize_content(self.metas['content'])
		
		
		for meta in self.metas:
			if isinstance(self.metas[meta],float):
				self.metas[meta] = str(round(self.metas[meta],2))
			else:
				if isinstance(self.metas[meta],basestring):
					try:
						self.metas[meta] = self.metas[meta].encode("utf-8")
					except UnicodeDecodeError:
						pass
			#~ print meta, self.metas[meta]
			
		
			
		self.db.save_data(id_product, self.metas, self.id_task)
		self.upload_images()
		
		return 1

	def run(self):
		"""start complete crawler"""
		
		self.logger.info("[run] iniciando(Completo=%s)" % self.mode_complete)
		
		try:
			self.db.init_task(self.id_task)
			
			
			for url_discover in self.config['discover_urls']:
				page = 1
				n_products = 1
				self.last_first_id = None
				while n_products > 0:
					if page > 1 and not "%d" in url_discover:
						#~ print "Saliendo", url_discover
						break
					try:
						url = url_discover % page
					except TypeError:
						url = url_discover
						
					self.logger.info("[run] recorriendo %s" % url)
					
					if "campana" in url:
						n_products = self.extract_product_campana(url)
					else:
						n_products = self.extract_product(url)
					page += 1
					#~ if page > 2: break;
					self.logger.info("[run] extraidos %d productos de %s" % (n_products, url))
					
					
			self.generate_csv()
			
			self.db.finish_task(self.id_task)
		except Exception as e:
			self.db.finish_task(self.id_task, True)
			
			exc_type, exc_obj, exc_tb = sys.exc_info()

			self.logger.error("%s\n %d: %s" %(traceback.format_exc(), exc_tb.tb_lineno, str(e)))
			raise
		

		
if __name__ == '__main__':
	
	
	if len(sys.argv) == 1:
		crawl = CrawlerComics_2()
		crawl.run()
	else:
		if "http" in sys.argv[1]:
			for url in sys.argv[1:]:
				crawl = CrawlerComics_2()
				crawl.extract_product(url)
				crawl.generate_csv()
			
				crawl.db.finish_task(crawl.id_task)
		else:
			crawl = CrawlerComics_2(id_task = sys.argv[1], mode = sys.argv[2])
			crawl.run()
